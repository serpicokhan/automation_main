'''
 fmt = getattr(settings, 'LOG_FORMAT', None)
 lvl = getattr(settings, 'LOG_LEVEL', logging.DEBUG)

 logging.basicConfig(format=fmt, level=lvl)
 logging.debug(nepartbject.OrderId.id)
 '''
from django.shortcuts import render
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.db.models import Sum
import jdatetime
from django.http import HttpResponseRedirect
from django.http import HttpResponse
from django.views.decorators import csrf
import django.core.serializers
import logging
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from myapp.models.parts import *
#from django.core import serializers
import json
from django.forms.models import model_to_dict
from myapp.forms import PartForm
from django.urls import reverse_lazy
from django.db import transaction
from myapp.business.partutility import *
from django.contrib.auth.decorators import login_required
from django.contrib.auth.context_processors import PermWrapper
from django.db.models import Sum
from rest_framework.decorators import api_view

from rest_framework.response import Response
from openpyxl import load_workbook
import csv

@login_required
def list_part(request,id=None):
    #
    books=[]

    
    books = Part.objects.all().order_by('partName')
    #paging

    wos=PartUtility.doPaging(request,books)
    return render(request, 'myapp/part/partList.html', {'part': wos,'section':'list_part'})


##########################################################

def save_part_form(request, form, template_name,id=None):


    data = dict()
    if (request.method == 'POST'):
        if form.is_valid():
            form.save()
            data['form_is_valid'] = True
            books = Part.objects.all().order_by('partName')
            wos=PartUtility.doPaging(request,list(books))
            data['html_part_list'] = render_to_string('myapp/part/partialPartList.html', {
                'part': wos,
                'perms': PermWrapper(request.user)

            })
        else:
            data['form_is_valid'] = False

    context = {'form': form,'lId':id}


    data['html_part_form'] = render_to_string(template_name, context, request=request)
    return JsonResponse(data)
##########################################################
def save2_part_form(request, form, template_name,id=None):
    data = dict()
    try:



        if (request.method == 'POST'):
            if form.is_valid():
                form.save()
                data['form_is_valid'] = True
                data['partName']=form.instance.partName
                data['partId']=form.instance.id
            else:
                data['form_is_valid'] = False

        context = {'form': form,'lId':id}


        data['html_part_form'] = render_to_string(template_name, context, request=request)
    except Exception:
        print("error")


    return JsonResponse(data)
##########################################################


def part_delete(request, id):
    comp1 = get_object_or_404(Part, id=id)
    data = dict()
    if (request.method == 'POST'):
        comp1.delete()
        data['form_is_valid'] = True  # This is just to play along with the existing code
        companies =  Part.objects.all().order_by('partName')
        wos=PartUtility.doPaging(request,list(companies))
        #Tasks.objects.filter(partId=id).update(part=id)
        data['html_part_list'] = render_to_string('myapp/part/partialPartList.html', {
            'part': wos,
            'perms': PermWrapper(request.user)
        })
    else:
        context = {'part': comp1}
        data['html_part_form'] = render_to_string('myapp/part/partialPartDelete.html',
            context,
            request=request,
        )
    return JsonResponse(data)

##########################################################

##########################################################
def part_create(request):
    if (request.method == 'POST'):
        form = PartForm(request.POST)
        return save_part_form(request, form, 'myapp/part/partialPartCreate.html')
    else:
        # partInstance=Part.objects.create()
        form = PartForm()
        return save_part_form(request, form, 'myapp/part/partialPartCreate.html')

@csrf_exempt
def part_create2(request):
    if (request.method == 'POST'):
        body_unicode = request.body.decode('utf-8')
        body = json.loads(body_unicode)
        data = request.POST.dict()
        data['partName']=body['partName']
        data['partDescription']=body['partDescription']
        data['partCode']=body['partCode']
        # data['partNote']=body['partNote']

        form = PartForm2(data=data)

        return save2_part_form(request, form, 'myapp/part/partialPartCreate2.html')
    else:
        # partInstance=Part.objects.create()
        form = PartForm2()
        return save2_part_form(request, form, 'myapp/part/partialPartCreate2.html')




##########################################################
def part_update(request, id):
    company= get_object_or_404(Part, id=id)
    template=""
    if (request.method == 'POST'):
        form = PartForm(request.POST, instance=company)
    else:
        assetcatText=company.partCategory.name if company.partCategory else ''
        form = PartForm(instance=company,initial={'partcategorytxt':assetcatText})
    fmt = getattr(settings, 'LOG_FORMAT', None)
    lvl = getattr(settings, 'LOG_LEVEL', logging.DEBUG)

    logging.basicConfig(format=fmt, level=lvl)
    logging.debug(id)

    return save_part_form(request, form,"myapp/part/partialPartUpdate.html",id)
##########################################################

##########################################################
#######################Search By tags#####################
def part_searchPart(request,searchStr):
    data=dict()

    searchStr=searchStr.replace('_',' ')
    books=None
    if(len(searchStr)==0):
        books=Parts.objects.all().order_by('partName')
    else:
        books=PartUtility.seachPart(searchStr).order_by('partName')
    wos=PartUtility.doPaging(request,list(books))
    data['html_part_search_tag_list'] = render_to_string('myapp/part/partialPartList.html', {               'part': wos  ,'perms': PermWrapper(request.user)                       })
    data['html_part_paginator'] = render_to_string('myapp/part/partialWoPagination.html', {
          'wo': wos,'pageType':'part_searchPart','ptr':searchStr})
    data['form_is_valid'] = True
    return JsonResponse(data)
#######################Search#####################
@csrf_exempt
def partCancel(request,id):
    try:

        data=dict()
        if(request.method=='POST'):
            tg=Part.objects.get(id=id)
            if(tg):
                if(not tg.partName or not tg.partCode):
                    tg.delete()
                    data['form_is_valid'] = True  # This is just to play along with the existing code
                    companies =  Part.objects.all().order_by('partName')
                    wos=PartUtility.doPaging(request,companies)
                    #Tasks.objects.filter(taskGroupId=id).update(taskGroup=id)
                    data['html_part_list'] = render_to_string('myapp/part/partialPartList.html', {
                        'part': wos
                    })

        return JsonResponse(data)
    except:
        return JsonResponse(dict())
####################################################
def inventoryLevel(request,id):
    data=dict()
    p=Stock.objects.filter(stockItem=id)
    data['part_inventory_level'] = render_to_string('myapp/part/partialinventorylevel.html', {
        'wos': p
    })
    return JsonResponse(data)
def inventorySum(request,id):
    data=dict()
    p=Stock.objects.filter(stockItem=id).aggregate(Sum('qtyOnHand'))
    data['part_inventory_sum'] = p
    return JsonResponse(data)
def partUsage(request,id):
    data=dict()
    n1=PartUtility.getPartMaintenancePie(id)
    n2=PartUtility.getWoPartUsageHistory(id)
    n3=PartUtility.getPartPurchaseHistory(id)

    s1,s2=[],[]
    z1,z2=[],[]
    x1,x2=[],[]
    for i in n1:
        s1.append('{:.2f}'.format(i.id))
        s2.append(i.mname)
    for i in n2:
        z1.append('{:.2f}'.format(i.id))
        z2.append(i.month)
    for i in n3:
        x1.append('{:.2f}'.format(i.id))
        x2.append(i.month)
    data['html_part_maintenance_type_list'] ={'woCompletedNum': s1,
                                    'woCompletedAssetId':s2,
                                    'lineAssetofflinecount':z1,'lineminthname':z2,
                                    'linepartpurchasecount':x1,'linepartpurchasemonth':x2

                                    }
    return JsonResponse(data)
def getPartConsumedItem(request,id,num):
    data=dict()
    wos=PartUtility.getConsumeInfo(id,num)
    data['form_is_valid']=True
    # data['html_stock_page']=render_to_string('myapp/stock/partialStocklist.html', {       '': q      })
    data['html_stock_list'] =render_to_string('myapp/part/consumedpartresult.html', {
        'wos': wos
    })
    return JsonResponse(data)
### get stock purchased history item in wo in stockslistdetail.html
def getPartPurchasedItem(request,id,num):
    data=dict()
    wos=PartUtility.getPurchasedInfo(id,num)
    data['form_is_valid']=True
    # data['html_stock_page']=render_to_string('myapp/stock/partialStocklist.html', {       '': q      })
    data['html_stock_list'] =render_to_string('myapp/part/purchasepartresult.html', {
        'wos': wos
    })
    return JsonResponse(data)

##########################################
def get_partCategory(request):
    data=dict()
    '''render_to_string('myapp/asset/temp.txt')'''
    m=PartUtility.getCategory()

    m=m.replace('"',"'")
    data["modalassetcat"]=render_to_string('myapp/part/partcategoryselector.html',{'cat':m,'perms': PermWrapper(request.user)})
    return JsonResponse(data)
##########################################
@api_view(['GET'])
def part_collection(request):
    if request.method == 'GET':
        # print("!23")
        posts = Part.objects.all()
        serializer = PartSerializer(posts, many=True)

        return Response(serializer.data)
@api_view(['GET'])
def part_detail_collection(request,id):
    if request.method == 'GET':
        # print("!23")
        posts = Part.objects.get(id=id)
        serializer = PartSerializer(posts)

        return Response(serializer.data)
##############################
###################################################################

def get_parts(request):
    searchStr= request.GET['qry'] if request.GET['qry'] else ''
    x=list(PartUtility.getParts(searchStr))
    # if(len(x)==0):
    #     print("dasdsa")
    #     x=[{'id':-1,'partName':'قطعه یافت نشد'}]


    # response_data = {}
    # response_data['result'] = '[dsadas,dasdasdas]'
    return JsonResponse(x, safe=False)
def col_letter_to_index(letter):
    index = 0
    for char in letter:
        index = index * 26 + (ord(char.upper()) - ord('A')) + 1
    return index - 1  # Subtract 1 to make it 0-based index
def upload_part(request):
    return render(request, 'myapp/part/partUpload.html', {})
def upload_file_part(request):
    def iter_rows(ws):
        for row in ws.iter_rows():
            yield [cell.value for cell in row]
    if request.method == 'POST':
        my_file=request.FILES.get('file')
        msg=PartCsvFile.objects.create(msgFile=my_file)
        workbook = load_workbook(filename='media/'+msg.msgFile.name)
        # ws = workbook.active
        # item=Part(pk=None)
        # for i in list(iter_rows(ws)):


        #     if(i[19]!=None):#id
        #         item=Part(pk=None)
        #         item.partName=i[18]
        #         item.partDescription=i[14] if(i[14]) else '-'
        #         item.partCode=0
        #         item.save()
        # sheet = workbook['Sheet1']  # Replace 'Sheet1' with your sheet name
        

        # Or use the default sheet (usually the first one)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
                if(row[col_letter_to_index('i')] is not None and row[col_letter_to_index('x')] is not None):
                    item=Part(pk=None)
                    item.partName=row[col_letter_to_index('i')]
                    item.partDescription=row[col_letter_to_index('i')]
                    item.partCode=row[col_letter_to_index('e')]
                    item.save()
                    item2=Part(pk=None)
                    item2.partName=row[col_letter_to_index('x')]
                    item2.partDescription=row[col_letter_to_index('x')]
                    item2.partCode=row[col_letter_to_index('v')]
                    item2.save()
                    # print(row[col_letter_to_index('i')],row[col_letter_to_index('x')])


        data=dict()
        data["id"]=msg.id
        return JsonResponse(data)
        # return HttpResponse('')
    return JsonResponse({'post':'fasle'})
