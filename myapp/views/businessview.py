'''
 fmt = getattr(settings, 'LOG_FORMAT', None)
 lvl = getattr(settings, 'LOG_LEVEL', logging.DEBUG)

 logging.basicConfig(format=fmt, level=lvl)
 logging.debug(nebusinessbject.OrderId.id)
 '''
from django.shortcuts import render
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.template.loader import render_to_string
from django.db.models import Sum
import jdatetime
from django.http import HttpResponseRedirect
from django.http import HttpResponse
from django.views.decorators import csrf
import django.core.serializers
import logging
from django.conf import settings

from myapp.models.business import *
#from django.core import serializers
import json
from django.forms.models import model_to_dict
from myapp.forms import BusinessForm,MiniBusinessForm
from django.urls import reverse_lazy
from django.db import transaction
from django.contrib.auth.context_processors import PermWrapper
from myapp.business.BusiUtil import *
from openpyxl import load_workbook



def list_business(request,id=None):
    #
    books = Business.objects.all().order_by('name')

    wos=BusinessUtility.doPaging(request,books)

    return render(request, 'myapp/business/businessList.html', {'business': wos,'section':'list_business'})


##########################################################

def save_business_form(request, form, template_name,id=None):


    data = dict()
    if (request.method == 'POST'):
        if form.is_valid():
            form.save()
            data['form_is_valid'] = True
            books = Business.objects.all().order_by('name')
            wos=BusinessUtility.doPaging(request,books)
            data['html_business_list'] = render_to_string('myapp/business/partialBusinessList.html', {
                'business': wos
            })
        else:
            data['form_is_valid'] = False

    context = {'form': form,'lId':id}


    data['html_business_form'] = render_to_string(template_name, context, request=request)
    return JsonResponse(data)
##########################################################


def business_delete(request, id):
    comp1 = get_object_or_404(Business, id=id)
    data = dict()
    if (request.method == 'POST'):
        comp1.delete()
        data['form_is_valid'] = True  # This is just to play along with the existing code
        companies =  Business.objects.all().order_by('name')
        #Tasks.objects.filter(businessId=id).update(business=id)
        data['html_business_list'] = render_to_string('myapp/business/partialBusinessList.html', {
            'business': companies
        })
    else:
        context = {'business': comp1}
        data['html_business_form'] = render_to_string('myapp/business/partialBusinessDelete.html',
            context,
            request=request,
        )
    return JsonResponse(data)

##########################################################

##########################################################
def business_create(request):
    if (request.method == 'POST'):
        form = BusinessForm(request.POST)
        return save_business_form(request, form, 'myapp/business/partialBusinessCreate.html')
    else:
        businessInstance=Business.objects.create()
        form = BusinessForm()
        return save_business_form(request, form, 'myapp/business/partialBusinessCreate.html',businessInstance.id)




##########################################################
def business_update(request, id):
    company= get_object_or_404(Business, id=id)
    template=""
    if (request.method == 'POST'):
        form = BusinessForm(request.POST, instance=company)
    else:
        form = BusinessForm(instance=company)
    fmt = getattr(settings, 'LOG_FORMAT', None)
    lvl = getattr(settings, 'LOG_LEVEL', logging.DEBUG)

    logging.basicConfig(format=fmt, level=lvl)
    logging.debug(id)

    return save_business_form(request, form,"myapp/business/partialBusinessUpdate.html",id)
##########################################################

##########################################################
def businessCancel(request,id):
    data=dict()
    # tg=Business.objects.get(id=id)
    # if(tg):
    #     if(not tg.name):
    #         tg.delete()
    #         data['form_is_valid'] = True  # This is just to play along with the existing code
    #         companies =  Business.objects.all().order_by('name')
    #         #Tasks.objects.filter(taskGroupId=id).update(taskGroup=id)
    #         data['html_business_list'] = render_to_string('myapp/business/partialBusinessList.html', {
    #             'business': companies
    #         })

    return JsonResponse(data)
#####################
def business_search(request):
    data=dict()

    searchStr=request.GET.get("q","")
    books=BusinessUtility.seachBusiness(searchStr).order_by('name')
    wos=BusinessUtility.doPaging(request,list(books))
    data['html_business_search_tag_list'] = render_to_string('myapp/business/partialBusinessList.html', {               'business': wos  ,'perms': PermWrapper(request.user)                       })
    # data['html_business_paginator'] = render_to_string('myapp/business/partialBusinessPagination.html', {
    #       'business': wos,'pageType':'business_search','ptr':searchStr})
    data['form_is_valid'] = True
    return JsonResponse(data)
def col_letter_to_index(letter):
    index = 0
    for char in letter:
        index = index * 26 + (ord(char.upper()) - ord('A')) + 1
def upload_business(request):
    return render(request, 'myapp/business/businessUpload.html', {})
def upload_file_business(request):
    def iter_rows(ws):
        for row in ws.iter_rows():
            yield [cell.value for cell in row]
    if request.method == 'POST':
        # print("here!!!",request.FILES)
        my_file=request.FILES.get('file')
        # # print(my_file.name)
        # month=request.POST.get('month_select',False)
        # sal=request.POST.get('sal_select',False)




        msg=BusinessCsvFile.objects.create(msgFile=my_file)

        print("@!!!!!!!!!!!!!!!")
        workbook = load_workbook(filename='media/'+msg.msgFile.name)
        ws = workbook.active
        # print(list(iter_rows(ws))[1])
        # Part.objects.filter(id__gt=20).delete()
       
        # item_old=Part(pk=None)
        #
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
                if(row[0] is not None and row[1] is not None):
                    item=Business(pk=None)
                    item.name=row[0]
                    item.code=row[1]                   
                    item.save()





        data=dict()
        data["id"]=msg.id
        return JsonResponse(data)
        # return HttpResponse('')
    return JsonResponse({'post':'fasle'})
def new_supplier(request):
    data=dict()
    sup=request.GET.get("qry","")
    if(len(sup)>0):
        instance=Business.objects.create(name=sup)
        data['is_valid']=True
        data['instance']=instance.id
    return JsonResponse(data)
def get_suppliers(request):
    searchStr= request.GET['qry'] if request.GET['qry'] else ''
    x=list(BusinessUtility.gets(searchStr))
    # if(len(x)==0):
    #     Business.create(name=searchStr)
    #     x=list(BusinessUtility.gets(searchStr))


    # if(len(x)==0):
    #     print("dasdsa")
    #     x=[{'id':-1,'partName':'قطعه یافت نشد'}]


    # response_data = {}
    # response_data['result'] = '[dsadas,dasdasdas]'
    return JsonResponse(x, safe=False)
